from __future__ import annotations

from pathlib import Path

import pandas as pd


class ExcelReportWriter:
    """Writes a clean monthly expense report workbook."""

    TRANSACTION_COLUMNS = [
        "Transaction Date",
        "Post Date",
        "Merchant",
        "Location",
        "Transaction ID",
        "Card Last 4",
        "Amount",
        "Category",
        "Raw Description",
    ]

    def write(self, transactions: pd.DataFrame, output_path: str | Path) -> None:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        transactions = self._ordered_transactions(transactions)
        summary = self._build_summary(transactions)
        metrics = self._build_metrics(transactions)

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            transactions.to_excel(writer, sheet_name="Transactions", index=False)
            summary.to_excel(writer, sheet_name="Summary", index=False, startrow=0, startcol=0)
            metrics.to_excel(writer, sheet_name="Summary", index=False, startrow=0, startcol=4)

            self._format_workbook(writer, transactions, summary, metrics)

    def _ordered_transactions(self, transactions: pd.DataFrame) -> pd.DataFrame:
        df = transactions.copy()
        for col in self.TRANSACTION_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[self.TRANSACTION_COLUMNS]

    def _build_summary(self, transactions: pd.DataFrame) -> pd.DataFrame:
        summary = (
            transactions.groupby("Category", dropna=False, as_index=False)
            .agg(Count=("Amount", "count"), Total_Amount=("Amount", "sum"))
            .sort_values("Total_Amount", ascending=False)
            .reset_index(drop=True)
        )

        grand_total = pd.DataFrame(
            [{
                "Category": "Grand Total",
                "Count": int(summary["Count"].sum()),
                "Total_Amount": float(summary["Total_Amount"].sum()),
            }]
        )
        return pd.concat([summary, grand_total], ignore_index=True)

    def _build_metrics(self, transactions: pd.DataFrame) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Metric": [
                    "Total Spend",
                    "Transactions",
                    "Average Transaction",
                    "Largest Transaction",
                ],
                "Value": [
                    float(transactions["Amount"].sum()),
                    int(len(transactions)),
                    float(transactions["Amount"].mean()) if len(transactions) else 0.0,
                    float(transactions["Amount"].max()) if len(transactions) else 0.0,
                ],
            }
        )

    def _format_workbook(
        self,
        writer: pd.ExcelWriter,
        transactions: pd.DataFrame,
        summary: pd.DataFrame,
        metrics: pd.DataFrame,
    ) -> None:
        workbook = writer.book

        money_format = "$#,##0.00"
        header_fill = "1F4E78"
        header_font = "FFFFFF"

        for sheet_name in ["Transactions", "Summary"]:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"

            for cell in ws[1]:
                cell.fill = self._fill(workbook, header_fill)
                cell.font = self._font(workbook, header_font, bold=True)

        tx_ws = writer.sheets["Transactions"]
        tx_ws.auto_filter.ref = tx_ws.dimensions
        tx_ws.column_dimensions["A"].width = 18
        tx_ws.column_dimensions["B"].width = 12
        tx_ws.column_dimensions["C"].width = 30
        tx_ws.column_dimensions["D"].width = 24
        tx_ws.column_dimensions["E"].width = 16
        tx_ws.column_dimensions["F"].width = 12
        tx_ws.column_dimensions["G"].width = 12
        tx_ws.column_dimensions["H"].width = 22
        tx_ws.column_dimensions["I"].width = 55

        amount_col_idx = list(transactions.columns).index("Amount") + 1
        for row in range(2, len(transactions) + 2):
            tx_ws.cell(row=row, column=amount_col_idx).number_format = money_format

        summary_ws = writer.sheets["Summary"]
        summary_ws.column_dimensions["A"].width = 24
        summary_ws.column_dimensions["B"].width = 12
        summary_ws.column_dimensions["C"].width = 16
        summary_ws.column_dimensions["E"].width = 24
        summary_ws.column_dimensions["F"].width = 18

        for row in range(2, len(summary) + 2):
            summary_ws.cell(row=row, column=3).number_format = money_format

        for row in range(2, len(metrics) + 2):
            metric_name = str(summary_ws.cell(row=row, column=5).value)
            if metric_name != "Transactions":
                summary_ws.cell(row=row, column=6).number_format = money_format

    def _fill(self, workbook, color: str):
        from openpyxl.styles import PatternFill

        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _font(self, workbook, color: str, bold: bool = False):
        from openpyxl.styles import Font

        return Font(color=color, bold=bold)
